from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from sqlalchemy.orm import Session
from typing import List, Optional
from pydantic import BaseModel
from datetime import datetime
from database import get_db
from models.layout import Layout
import os
import uuid
import json
from pathlib import Path

router = APIRouter(
    prefix="/api/templates",
    tags=["templates"]
)

# テンプレートディレクトリの設定
TEMPLATE_DIR = Path("templates_excel")
TEMPLATE_DIR.mkdir(exist_ok=True)

# Pydanticモデル
class LayoutCreate(BaseModel):
    name: str
    description: Optional[str] = None
    default_width: Optional[int] = 400
    default_height: Optional[int] = 300

class LayoutResponse(BaseModel):
    id: int
    name: str
    description: Optional[str]
    excel_filename: Optional[str]
    excel_file_path: Optional[str]
    cell_data: Optional[str]
    layout_config: Optional[str]
    cell_styles: Optional[str]
    merged_cells: Optional[str]
    default_width: int
    default_height: int
    created_at: datetime
    updated_at: datetime

    class Config:
        from_attributes = True

@router.get("/", response_model=List[LayoutResponse])
def get_templates(skip: int = 0, limit: int = 100, db: Session = Depends(get_db)):
    """全テンプレートの取得"""
    templates = db.query(Layout).offset(skip).limit(limit).all()
    return templates

@router.get("/{template_id}", response_model=LayoutResponse)
def get_template(template_id: int, db: Session = Depends(get_db)):
    """特定テンプレートの取得"""
    template = db.query(Layout).filter(Layout.id == template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail="テンプレートが見つかりません")
    return template

@router.get("/{template_id}/debug")
def get_template_debug(template_id: int, db: Session = Depends(get_db)):
    """テンプレートのデバッグ情報取得"""
    template = db.query(Layout).filter(Layout.id == template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail="テンプレートが見つかりません")
    
    # デバッグ情報を追加
    debug_info = {
        "template_id": template.id,
        "name": template.name,
        "excel_filename": template.excel_filename,
        "excel_file_path": template.excel_file_path,
        "file_exists": os.path.exists(template.excel_file_path) if template.excel_file_path else False,
        "has_cell_data": template.cell_data is not None,
        "has_layout_config": template.layout_config is not None,
        "has_cell_styles": template.cell_styles is not None,
        "has_merged_cells": template.merged_cells is not None,
        "cell_data_length": len(template.cell_data) if template.cell_data else 0,
        "layout_config_length": len(template.layout_config) if template.layout_config else 0,
        "cell_styles_length": len(template.cell_styles) if template.cell_styles else 0,
        "merged_cells_length": len(template.merged_cells) if template.merged_cells else 0,
        "default_width": template.default_width,
        "default_height": template.default_height,
        "created_at": template.created_at,
        "updated_at": template.updated_at
    }
    
    # セルデータの内容をサンプルとして取得
    if template.cell_data:
        try:
            cell_data_parsed = json.loads(template.cell_data)
            debug_info["cell_data_sample"] = dict(list(cell_data_parsed.items())[:5])
        except:
            debug_info["cell_data_sample"] = "JSON parse error"
    
    return {
        "debug_info": debug_info,
        "template": template
    }

@router.post("/", response_model=LayoutResponse)
def create_template(template: LayoutCreate, db: Session = Depends(get_db)):
    """新規テンプレートの作成"""
    db_template = Layout(**template.dict())
    db.add(db_template)
    db.commit()
    db.refresh(db_template)
    return db_template

@router.post("/upload", response_model=LayoutResponse)
async def upload_excel_template(
    file: UploadFile = File(...),
    name: str = Form(...),
    description: Optional[str] = Form(None),
    db: Session = Depends(get_db)
):
    """Excelテンプレートのアップロード"""
    # ファイル形式の確認
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Excelファイルのみアップロード可能です")
    
    # ファイル名を生成
    file_extension = Path(file.filename).suffix
    unique_filename = f"{uuid.uuid4()}{file_extension}"
    
    # ファイルを保存
    file_path = TEMPLATE_DIR / unique_filename
    with open(file_path, "wb") as buffer:
        content = await file.read()
        buffer.write(content)
    
    # Excelファイルを解析
    try:
        excel_data = analyze_excel_file(str(file_path))
    except Exception as e:
        # ファイルを削除
        if os.path.exists(file_path):
            os.remove(file_path)
        raise HTTPException(status_code=400, detail=f"Excelファイルの解析に失敗しました: {str(e)}")
    
    # データベースに登録
    db_template = Layout(
        name=name,
        description=description,
        excel_filename=file.filename,
        excel_file_path=str(file_path),
        cell_data=excel_data.get('cell_data'),
        merged_cells=excel_data.get('merged_cells'),
        cell_styles=excel_data.get('cell_styles'),
        layout_config=excel_data.get('layout_config'),
        default_width=excel_data.get('default_width', 400),
        default_height=excel_data.get('default_height', 300)
    )
    
    db.add(db_template)
    db.commit()
    db.refresh(db_template)
    
    return db_template

def analyze_excel_file(file_path: str) -> dict:
    """Excelファイルを解析してレイアウト情報を取得"""
    try:
        import openpyxl
        from openpyxl.styles import Border, Side
        
        workbook = openpyxl.load_workbook(file_path)
        worksheet = workbook.active
        
        # セルデータを取得
        cell_data = {}
        cell_styles = {}
        max_row = 0
        max_col = 0
        
        # 使用されているセル範囲を取得
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None or has_border(cell):
                    max_row = max(max_row, cell.row)
                    max_col = max(max_col, cell.column)
                    
                    cell_address = f"{cell.column_letter}{cell.row}"
                    
                    # セルの値を取得
                    if cell.value is not None:
                        cell_data[cell_address] = str(cell.value)
                    
                    # セルのスタイル情報を取得
                    cell_styles[cell_address] = {
                        'border': get_border_info(cell),
                        'fill': get_fill_info(cell),
                        'font': get_font_info(cell),
                        'alignment': get_alignment_info(cell)
                    }
        
        # 結合セル情報を取得
        merged_cells = []
        for merged_range in worksheet.merged_cells.ranges:
            merged_cells.append({
                'start': str(merged_range.min_col) + str(merged_range.min_row),
                'end': str(merged_range.max_col) + str(merged_range.max_row),
                'range': str(merged_range)
            })
        
        # 行・列のサイズ情報を取得
        row_heights = {}
        col_widths = {}
        
        for row_num in range(1, max_row + 1):
            if worksheet.row_dimensions[row_num].height:
                row_heights[row_num] = worksheet.row_dimensions[row_num].height
        
        for col_num in range(1, max_col + 1):
            col_letter = openpyxl.utils.get_column_letter(col_num)
            if worksheet.column_dimensions[col_letter].width:
                col_widths[col_letter] = worksheet.column_dimensions[col_letter].width
        
        # レイアウト設定
        layout_config = {
            'max_row': max_row,
            'max_col': max_col,
            'row_heights': row_heights,
            'col_widths': col_widths,
            'sheet_name': worksheet.title
        }
        
        # 推奨サイズを計算
        default_width = min(max(sum(col_widths.values()) * 8, 300), 600)
        default_height = min(max(sum(row_heights.values()) * 1.2, 200), 500)
        
        return {
            'cell_data': json.dumps(cell_data, ensure_ascii=False),
            'merged_cells': json.dumps(merged_cells, ensure_ascii=False),
            'cell_styles': json.dumps(cell_styles, ensure_ascii=False),
            'layout_config': json.dumps(layout_config, ensure_ascii=False),
            'default_width': int(default_width),
            'default_height': int(default_height)
        }
        
    except ImportError:
        raise Exception("openpyxlライブラリがインストールされていません")
    except Exception as e:
        raise Exception(f"Excelファイルの解析中にエラーが発生しました: {str(e)}")

def has_border(cell) -> bool:
    """セルに罫線があるかチェック"""
    border = cell.border
    return any([
        border.left.style,
        border.right.style,
        border.top.style,
        border.bottom.style
    ])

def get_border_info(cell) -> dict:
    """セルの罫線情報を取得"""
    border = cell.border
    return {
        'left': {'style': border.left.style, 'color': str(border.left.color.rgb) if border.left.color else None},
        'right': {'style': border.right.style, 'color': str(border.right.color.rgb) if border.right.color else None},
        'top': {'style': border.top.style, 'color': str(border.top.color.rgb) if border.top.color else None},
        'bottom': {'style': border.bottom.style, 'color': str(border.bottom.color.rgb) if border.bottom.color else None}
    }

def get_fill_info(cell) -> dict:
    """セルの塗りつぶし情報を取得"""
    fill = cell.fill
    return {
        'type': fill.fill_type,
        'color': str(fill.start_color.rgb) if fill.start_color else None
    }

def get_font_info(cell) -> dict:
    """セルのフォント情報を取得"""
    font = cell.font
    return {
        'name': font.name,
        'size': font.size,
        'bold': font.bold,
        'italic': font.italic,
        'color': str(font.color.rgb) if font.color else None
    }

def get_alignment_info(cell) -> dict:
    """セルの配置情報を取得"""
    alignment = cell.alignment
    return {
        'horizontal': alignment.horizontal,
        'vertical': alignment.vertical,
        'wrap_text': alignment.wrap_text
    }

@router.put("/{template_id}", response_model=LayoutResponse)
def update_template(template_id: int, template: LayoutCreate, db: Session = Depends(get_db)):
    """テンプレートの更新"""
    db_template = db.query(Layout).filter(Layout.id == template_id).first()
    if not db_template:
        raise HTTPException(status_code=404, detail="テンプレートが見つかりません")
    
    for key, value in template.dict().items():
        setattr(db_template, key, value)
    
    db_template.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(db_template)
    return db_template

@router.post("/create-web", response_model=LayoutResponse)
def create_web_template(template_data: dict, db: Session = Depends(get_db)):
    """Webエディタで作成されたテンプレートの保存"""
    try:
        # セルデータを変換
        cell_data = {}
        if template_data.get('cell_data'):
            for cell_id, data in template_data['cell_data'].items():
                if data.get('text'):
                    # セル座標をExcel形式に変換 (例: "0-1" -> "B1")
                    row, col = map(int, cell_id.split('-'))
                    excel_address = f"{chr(65 + col)}{row + 1}"
                    cell_data[excel_address] = data['text']
        
        # レイアウト設定を作成
        layout_config = {
            'max_row': template_data.get('rows', 8),
            'max_col': template_data.get('cols', 6),
            'row_heights': {},
            'col_widths': {},
            'sheet_name': 'WebTemplate'
        }
        
        # セルスタイルを変換
        cell_styles = {}
        if template_data.get('cell_styles'):
            for cell_id, style_data in template_data['cell_styles'].items():
                row, col = map(int, cell_id.split('-'))
                excel_address = f"{chr(65 + col)}{row + 1}"
                
                # CSSスタイルとクラス名をそのまま保存
                cell_style = {
                    'style': style_data.get('style', ''),
                    'className': style_data.get('className', ''),
                    'border': {'left': {}, 'right': {}, 'top': {}, 'bottom': {}},
                    'fill': {},
                    'font': {},
                    'alignment': {}
                }
                
                if style_data.get('style'):
                    css_style = style_data['style']
                    
                    # text-alignを特別に処理
                    if 'text-align:' in css_style:
                        text_align = css_style.split('text-align:')[1].split(';')[0].strip()
                        cell_style['alignment']['horizontal'] = text_align
                    
                    # 背景色
                    if 'background-color' in css_style:
                        color = css_style.split('background-color:')[1].split(';')[0].strip()
                        cell_style['fill']['color'] = color
                    
                    # 罫線
                    if 'border:' in css_style:
                        border_style = css_style.split('border:')[1].split(';')[0].strip()
                        cell_style['border']['left']['style'] = 'thin'
                        cell_style['border']['right']['style'] = 'thin'
                        cell_style['border']['top']['style'] = 'thin'
                        cell_style['border']['bottom']['style'] = 'thin'
                
                if style_data.get('className'):
                    class_name = style_data['className']
                    if 'text-bold' in class_name:
                        cell_style['font']['bold'] = True
                    if 'text-italic' in class_name:
                        cell_style['font']['italic'] = True
                    if 'text-center' in class_name:
                        cell_style['alignment']['horizontal'] = 'center'
                    if 'text-right' in class_name:
                        cell_style['alignment']['horizontal'] = 'right'
                
                cell_styles[excel_address] = cell_style
                
                # 同じスタイルを内部ID形式でも保存（カメラ画面での利用のため）
                cell_styles[cell_id] = cell_style
        
        # 結合セル情報を変換（クライアント形式を保持）
        merged_cells_raw = template_data.get('merged_cells', [])
        print(f"SERVER_DEBUG: 受信した結合データ: {merged_cells_raw}")
        
        # Map.entries()形式をそのまま保存（データ形式を変更しない）
        merged_cells = merged_cells_raw
        print(f"SERVER_DEBUG: 保存する結合データ（変換なし）: {merged_cells}")
        
        # セルタイプとセル設定データを処理
        cell_types = template_data.get('cell_types', {})
        cell_configs = template_data.get('cell_configs', {})
        cell_sizes = template_data.get('cell_sizes', {})
        
        # レイアウト設定にセルタイプとセル設定を追加
        layout_config['cell_types'] = cell_types
        layout_config['cell_configs'] = cell_configs
        layout_config['cell_sizes'] = cell_sizes
        
        # データベースに保存
        db_template = Layout(
            name=template_data['name'],
            description=template_data.get('description', ''),
            excel_filename=None,  # Webテンプレートなので null
            excel_file_path=None,
            cell_data=json.dumps(cell_data, ensure_ascii=False),
            merged_cells=json.dumps(merged_cells, ensure_ascii=False),
            cell_styles=json.dumps(cell_styles, ensure_ascii=False),
            layout_config=json.dumps(layout_config, ensure_ascii=False),
            default_width=template_data.get('default_width', 400),
            default_height=template_data.get('default_height', 300)
        )
        
        db.add(db_template)
        db.commit()
        db.refresh(db_template)
        
        return db_template
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"テンプレートの作成に失敗しました: {str(e)}")

@router.delete("/{template_id}")
def delete_template(template_id: int, db: Session = Depends(get_db)):
    """テンプレートの削除"""
    db_template = db.query(Layout).filter(Layout.id == template_id).first()
    if not db_template:
        raise HTTPException(status_code=404, detail="テンプレートが見つかりません")
    
    # Excelファイルを削除
    if db_template.excel_file_path and os.path.exists(db_template.excel_file_path):
        os.remove(db_template.excel_file_path)
    
    # データベースから削除
    db.delete(db_template)
    db.commit()
    
    return {"message": "テンプレートが削除されました"}